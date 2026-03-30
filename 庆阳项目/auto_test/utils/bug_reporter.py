# -*- coding: utf-8 -*-
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 与业务约定一致：接口名称 = 路径（如 /api/template/pageTemplate），不含域名
_HEADERS = ["用例编号", "接口名称", "URL", "请求方法", "请求参数", "失败原因", "时间戳"]
_COL_WIDTHS = [42, 36, 52, 10, 70, 55, 18]
_RUN_HEADERS = ["时间戳", "通过", "失败", "跳过", "xfail", "错误", "退出码", "说明"]


def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(s, limit=12000):
    s = str(s) if s is not None else ""
    return s[:limit] + ("...(截断)" if len(s) > limit else "")


def write_bug(
    path: str,
    case_id: str,
    interface_name: str,
    url: str,
    method: str,
    request_params: str,
    failure_reason: str,
    sheet: str = "AutoBugReport",
) -> None:
    """
    写入一条失败记录。列：用例编号、接口名称(API路径)、完整 URL、HTTP 方法、请求参数、失败原因、时间戳。
    """
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        hfill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        b = _border()
        for col, h in enumerate(_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.border = b
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        for i, w in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        ws = wb[sheet]
        # 若表头不是新版（例如旧文件），重写表头行，避免列错位
        if ws.max_row == 0 or ws.cell(1, 1).value != _HEADERS[0]:
            for col, h in enumerate(_HEADERS, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.border = _border()
            for i, w in enumerate(_COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    row = ws.max_row + 1
    data = [
        _cell(case_id, 800),
        _cell(interface_name, 500),
        _cell(url, 2500),
        _cell(method, 20),
        _cell(request_params, 12000),
        _cell(failure_reason, 12000),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    rfill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    b = _border()
    for col, v in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = b
        c.fill = rfill
        c.alignment = Alignment(vertical="top", wrap_text=True)
    try:
        nl = str(data[5]).count("\n") + 1
        ws.row_dimensions[row].height = min(180, max(36, 14 * nl))
    except Exception:
        ws.row_dimensions[row].height = 55

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


def append_run_log(path, passed, failed, skipped, xfailed, errors, exit_code, note=""):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    if "RunLog" not in wb.sheetnames:
        ws = wb.create_sheet("RunLog")
        hfill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        b = _border()
        for col, h in enumerate(_RUN_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.border = b
    else:
        ws = wb["RunLog"]
    ws = wb["RunLog"]
    r = ws.max_row + 1
    row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), passed, failed, skipped, xfailed, errors, exit_code,
           note or ("有失败见 AutoBugReport" if failed else "无失败")]
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") if failed == 0 else PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    b = _border()
    for col, v in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=v)
        c.border = b
        c.fill = fill
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
